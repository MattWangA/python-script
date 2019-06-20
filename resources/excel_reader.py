# coding: utf-8
import sys
from openpyxl import load_workbook, utils

class ExcelReader:
	_worksheet = None
	_last_column = None
	_last_row = None

	def __init__(self, file, last_column=None, last_row=None, worksheet=None):
		self.__set_active_worksheet(file, worksheet)
		self.__set_last_column(last_column)
		self.__set_last_row(last_row)
		
	def __set_last_column(self, last_column):
		if last_column:
			self._last_column = last_column
		else:
			self._last_column = self._worksheet.max_column

	def __set_last_row(self, last_row):
		if last_row:
			self._last_row = last_row
		else:
			self._last_row = self._worksheet.max_row

	def __set_active_worksheet(self, file, worksheet):
		try:
			wb = load_workbook(filename=file, read_only=True)
			if worksheet:
				self._worksheet = wb[worksheet]
			else:
				self._worksheet = wb.active
		except IOError:
			sys.exit('Cannot find find file {%s}' % file)
		except utils.exceptions.InvalidFileException:
			sys.exit('Please check file {%s}. Supported formats are: .xlsx,.xlsm,.xltx,.xltm' % file)

	def get_headers(self):
		return self.get_worksheet_data()[0]

	def get_active_worksheet(self):
		return self._worksheet

	def get_worksheet_data(self):
		data = []
		for row in self._worksheet.iter_rows(min_col = 1, max_col = self._last_column, min_row = 1, max_row = self._last_row):
			row_values = []
			for i in range(len(row)):
				if isinstance(row[i].value, str):
					row_values.append(row[i].value.strip())
				else:
					row_values.append(row[i].value)
			data.append(row_values)
		return data

	def get_worksheet_data_dict(self):
		data_list = self.get_worksheet_data()

		headers = data_list[0]
		final_list = []
		for row in data_list[1:]:
			temp_dict = {}
			for i in range(len(headers)):
				temp_dict[headers[i]] = row[i]
			if not all(value==None for value in list(temp_dict.values())):
				final_list.append(temp_dict)
		return final_list

def get_headers(file, last_column=None, last_row=None, worksheet=None):
	return ExcelReader(file, last_column, last_row, worksheet).get_headers()

def get_dict(file, last_column=None, last_row=None, worksheet=None):
	return ExcelReader(file, last_column, last_row, worksheet).get_worksheet_data_dict()

def get_list(file, last_column=None, last_row=None, worksheet=None):
	return ExcelReader(file, last_column, last_row, worksheet).get_worksheet_data()

#tests
def run_tests():
	test_error_when_file_does_not_exist()
	test_when_not_excel_file()
	test_when_excel_file()
	print ('Tests: Complete')

def test_error_when_file_does_not_exist():
	try:
		get_list('file_that_does_not_exist.txt')
		raise Exception('Fail: No error found for missing file.')
	except:
		print('Pass: Error when File does not exist.')

def test_when_not_excel_file():
	get_list('test.txt')

def test_when_excel_file():
	print((get_list('test.xlsx')))

def test_read_active_sheet_as_dict():
	print((get_dict('test.xlsx', last_column=3)))
